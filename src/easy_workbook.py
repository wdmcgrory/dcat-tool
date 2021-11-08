#!/usr/bin/env python3
#
# https://stackoverflow.com/questions/60824183/how-to-get-each-char-format-from-xlsx-cell-with-xlrd?
import datetime
import json
import os
import os.path
import sys

from openpyxl import Workbook
from io import BytesIO

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side, Protection, Font, Fill, Color
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_THICK
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
import openpyxl.styles.colors as colors

# https://htmlcolorcodes.com/

H1 = Font(bold=True, size=16)
H2 = Font(bold=True, size=14)
H3 = Font(bold=True, size=12)

BOLD        = Font(bold=True)
CENTERED    = Alignment(horizontal='center')
YELLOW_FILL = PatternFill(fill_type='solid', start_color='00ffff00', end_color='00ffff00')
PINK_FILL   = PatternFill(fill_type='solid', start_color='ffb6c1', end_color='ffb6c1')
LIGHT_GREEN_FILL  = PatternFill(fill_type='solid', start_color='EAFAF1', end_color='EAF8F1')

#PR_FILL = PatternFill(fill_type='solid', start_color='F4D03F', end_color='F4D03F')       # yellow
PR_FILL     = PatternFill(fill_type='solid', start_color='F4D03F')       # yellow
STATE1_FILL = PatternFill(fill_type='solid', start_color='F2F4F4')   # silver
STATE2_FILL = PatternFill(fill_type='solid', start_color='A9CCE3') # blue

LIGHT_COLORS = ['FFFFEE','FFEEFF','EEFFFF','FFEEEE','EEFFEE','EEEEFF','EEEEEE']

def darker(openpyxl_fill):
    rgb = openpyxl_fill.start_color.rgb
    (r,g,b) = [int(s,16) for s in [rgb[0:2],rgb[2:4],rgb[4:6]]]
    r = max(r-1,0)
    g = max(g-1,0)
    b = max(b-1,0)
    color = f"{r:02X}{g:02X}{b:02X}"
    return PatternFill(fill_type='solid', start_color=color)

class EasyWorkbook(Workbook):
    def __init__(self,*args,**kwargs):
        super().__init__(*args,**kwargs)

    def clean(self):
        """Remove default 'Sheet' if exists"""
        if 'Sheet' in self:
            del self['Sheet']

    def format_rows(self,ws,*,min_row,max_row,column=None,skip_blank=True,min_col,max_col,fills=None,value_fills=None,stripe=False):
        """Apply colors to each row when column changes. If column is not specified, change every row."""
        if fills:
            fills = RingBuffer(fills)
            fill  = fills.next()
        prev_fill  = None
        prev_value = None
        make_darker = False
        for cellrow in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            column_value = cellrow[column-min_col if column is not None else 0].value
            if skip_blank and column_value=='':
                continue

            make_darker  = not make_darker

            if not column_value:
                fill = None
            elif value_fills is not None and column_value in value_fills:
                fill = value_fills[column_value]
            elif column_value == prev_value:
                fill = prev_fill
                continue        #  don't change
            else:
                prev_fill   = fill       = fills.next()
                prev_value  = column_value
                make_darker = False

            #if make_darker and stripe:
            #    fill = darker(fill)

            for cell in cellrow:
                if fill:
                    cell.fill = fill

    def set_cell(ws,*,row,column,**kwargs):
        cell = ws.cell(row=row, column=column)
        for (key,value) in kwargs.items():
            setattr(cell, key, value)


class ColumnInfo:
    __slots__ = ('value', 'comment', 'author', 'width', 'typ', 'group')
    def __init__(self, **kwargs):
        for k,v in kwargs.items():
            setattr(self, k, v)
    def __str__(self):
        return f"<ColumnInfo value={self.value} comment={self.comment} width={self.width} typ={self.typ}>"

# Slightly higher-level
class ExcelGenerator:
    def __init__(self):
        self.wb = EasyWorkbook()
        self.wb.remove(self.wb.active)    # remove default sheet

    def add_markdown_sheet( self, name, markdown):
        ins = self.wb.create_sheet( name )
        for (row,line) in enumerate( markdown.split("\n"),1):
            font = None
            if line.startswith("# "):
                line = line[2:]
                font = H1
            if line.startswith("## "):
                line = line[3:]
                font = H2
            if line.startswith("### "):
                line = line[4:]
                font = H3
            ins.cell(row=row, column=1).value = line.strip()
            if font:
                ins.cell(row=row, column=1).font = font

    def add_columns_sheet(self, name, ci_list, rows=100):
        """ add (dcat name, display name, help, width, field type)"""
        if not isinstance(ci_list, list):
            raise ValueError(f"{ci_list} is not a list, it is a {type(ci_list)}")
        for ci in ci_list:
            if not isinstance(ci, ColumnInfo):
                raise ValueError(f"{ci} is not an instance of ColumnInfo, it is a {type(ci)}")
        ws = self.wb.create_sheet( name )

        # set up data validation
        # https://openpyxl.readthedocs.io/en/stable/validation.html

        dv_boolean = DataValidation(type="list", formula1='"true,false"', allow_blank=True)
        dv_boolean.error = 'Please select yes or no, or leave blank.'
        dv_boolean.errorTitle = 'Invalid Value'
        dv_boolean.prompt = 'Please select from the list'
        dv_boolean.promptTitle = 'List Selection'
        ws.add_data_validation(dv_boolean)

        dv_decimal = DataValidation(type='decimal')
        ws.add_data_validation(dv_decimal)

        dv_date    = DataValidation(type='date')
        ws.add_data_validation(dv_date)


        last_group = None
        color_index = 0
        my_border = Border(left=Side(style=BORDER_THIN),
                           right=Side(style=BORDER_THIN),
                           top=Side(style=BORDER_THIN),
                           bottom=Side(style=BORDER_THIN))

        group_border = Border(left=Side(style=BORDER_THICK),
                           right=Side(style=BORDER_THIN),
                           top=Side(style=BORDER_THIN),
                           bottom=Side(style=BORDER_THIN))

        for (col,obj) in enumerate( ci_list, 1):
            from openpyxl.comments import Comment
            import openpyxl.utils

#            print(obj.value, obj.typ)


            # We tried making the comment string the description and the DCATv3 type is the comment "author", but that didn't work
            cell = ws.cell(row=1, column=col)
            cell.value = obj.value
            cell.alignment = Alignment(textRotation=45)
            cell.comment = Comment(obj.comment + "\n" + "<" + obj.author + ">" , "")
            cell.comment.width = 200
            cell.comment.height = 400
            column_letter = openpyxl.utils.get_column_letter( cell.col_idx)
            ws.column_dimensions[ column_letter ].width   = obj.width

            new_border = my_border
            if last_group != obj.group:
                last_group = obj.group
                color_index = (color_index + 1) % len(LIGHT_COLORS)
                group_color = LIGHT_COLORS[color_index]
                color_fill = PatternFill(start_color = group_color, end_color=group_color, fill_type='solid')
                new_border = group_border
            cell.fill = color_fill
            cell.border = new_border
            for row in range(2,rows):
                cell2 = ws.cell(row=row, column=col)
                cell2.value=''
                cell2.fill = color_fill
                cell2.border = new_border

            # Set the types with data validation where possible
            if obj.typ=='xsd:boolean':
                dv_boolean.add(f'{column_letter}2:{column_letter}{rows}')


    def save(self, fname):
        self.wb.save(fname)

    def export(self):
        virtual_file = BytesIO()
        self.wb.save(virtual_file)
        return virtual_file.getvalue()
        

